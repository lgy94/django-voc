import openpyxl
from django.http import HttpResponse, request, HttpResponseRedirect
from django.shortcuts import render, get_object_or_404

# Create your views here.
from django.urls import reverse, reverse_lazy
from django.views.generic import CreateView, DetailView, UpdateView, DeleteView, ListView
from openpyxl.writer.excel import save_virtual_workbook

from categoryapp.models import Categorylist
from vocapp.forms import VocCreationForm
from vocapp.models import Voclist


def hello_world(request):
    return render(request, 'vocapp/hello_world.html')

class VocCreatView(CreateView):
    model = Voclist
    form_class = VocCreationForm
    context_object_name = 'voclist'
    success_url = reverse_lazy('categoryapp:list')
    template_name = 'vocapp/create.html'

class VocDetailView(DetailView):
    model= Voclist
    context_object_name = 'voclist'
    template_name = 'vocapp/detail.html'

class VocUpdateView(UpdateView):
    model = Voclist
    form_class = VocCreationForm
    success_url = reverse_lazy('vocapp:hello_world')
    template_name = 'vocapp/update.html'

class VocDeleteView(DeleteView):
    model = Voclist
    form_class = VocCreationForm
    success_url = reverse_lazy('vocapp:hello_world')
    template_name = 'vocapp/delete.html'

class VocListView(ListView):
    model = Voclist
    context_object_name = 'voclist'
    template_name = 'vocapp/list.html'

    def get_queryset(self):
        self.id = get_object_or_404(Categorylist, id=self.kwargs['pk'])
        return Voclist.objects.filter(category_num=self.id)

def downolad_excel(request):
    if request.method=="POST":
        category_num = request.POST.get('category_num')
        category_num = str(category_num)

        model = Voclist
        DBdata = list(model.objects.values().filter(category_num_id=category_num)) #category_num_id에 따른 엑셀값 뽑아내기
        Datalist=[]

        wb = openpyxl.Workbook()
        sheet= wb.active
        num = 0
        sheet.cell(row=1, column=1).value = '인입시간'
        sheet.cell(row=1, column=2).value = '인입방법'
        sheet.cell(row=1, column=3).value = '회원번호'
        sheet.cell(row=1, column=4).value = '회원명'
        sheet.cell(row=1, column=5).value = '문서번호'
        sheet.cell(row=1, column=6).value = '제목'
        sheet.cell(row=1, column=7).value = '내용'
        sheet.cell(row=1, column=8).value = '대응내역'
        sheet.cell(row=1, column=9).value = '담당자'
        sheet.cell(row=1, column=10).value = '장애보고서'
        sheet.cell(row=1, column=11).value = '파트너사'

        for list_data in DBdata:
            Datalist.append(list_data)            #배열을 통해 여러 열을 한번에 입력
            sheet.append([Datalist[num].get('voc_date'), Datalist[num].get('voc_method'), Datalist[num].get('client_number'), Datalist[num].get('client_name'), Datalist[num].get('voc_number'), Datalist[num].get('voc_title'), Datalist[num].get('voc_content'), Datalist[num].get('voc_comment'),Datalist[num].get('voc_manger'),Datalist[num].get('report'),Datalist[num].get('partner')])
            num+=1

        wb.save('excel.xlsx')
        wb.close()
        filename = 'excel_download'
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="' + filename + '.xlsx"'
        return response
